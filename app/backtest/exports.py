"""Backtest export helpers — CSV, Excel, JSON, PDF."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


TRADE_FIELDS = [
    "trade_num", "side", "entry_time", "exit_time", "entry_price", "exit_price",
    "price_move_pct", "pnl_usd", "bars_held", "exit_reason", "mfe_pct", "mae_pct",
]


def build_excel_bytes(row: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    ws.append(TRADE_FIELDS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for tr in row.get("trades", []):
        ws.append([tr.get(f) for f in TRADE_FIELDS])

    stats = row.get("statistics") or {}
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    for k, v in stats.items():
        ws2.append([k, v])

    monthly = row.get("monthly_report") or []
    if monthly:
        ws3 = wb.create_sheet("Monthly")
        ws3.append(["month", "trades", "profit", "win_rate", "profit_factor"])
        for m in monthly:
            ws3.append([m.get("month"), m.get("trades"), m.get("profit"), m.get("win_rate"), m.get("profit_factor")])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_bytes(row: dict[str, Any]) -> bytes:
    """Minimal text PDF summary (no external PDF lib)."""
    stats = row.get("statistics") or {}
    lines = [
        f"Backtest Run #{row.get('id', '')}",
        f"Strategy: {row.get('strategy_id', '')}",
        f"Symbol: {row.get('symbol', '')} · {row.get('timeframe', '')}",
        f"Period: {row.get('start_date', '')} → {row.get('end_date', '')}",
        "",
        "Results",
        f"  Total Return: {stats.get('total_return_pct')}%",
        f"  Net Profit: ${stats.get('net_profit')}",
        f"  Final Equity: ${stats.get('final_equity')}",
        f"  Profit Factor: {stats.get('profit_factor')}",
        f"  Win Rate: {stats.get('win_rate')}%",
        f"  Total Trades: {stats.get('total_trades')}",
        f"  Max Drawdown: {stats.get('max_drawdown_pct')}%",
        f"  Expectancy: ${stats.get('expectancy')}",
    ]
    text = "\n".join(lines)

    def _escape(s: str) -> str:
        return s.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    content = f"BT\n/F1 11 Tf\n50 750 Td\n14 TL\n({_escape(text)}) Tj\n"
    stream = content.encode("latin-1", errors="replace")
    parts = [
        b"%PDF-1.4\n",
        b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n",
        b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n",
        b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 4 0 R"
        b"/Resources<</Font<</F1 5 0 R>>>>>>endobj\n",
        f"4 0 obj<</Length {len(stream)}>>stream\n".encode() + stream + b"endstream\nendobj\n",
        b"5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj\n",
        b"xref\n0 6\n0000000000 65535 f \n",
    ]
    body = b"".join(parts)
    xref_offset = len(body)
    trailer = (
        f"trailer<</Size 6/Root 1 0 R>>\nstartxref\n{xref_offset}\n%%EOF\n"
    ).encode()
    return body + trailer
