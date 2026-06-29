"""API routes for SMA Signal Optimizer (research only)."""

from __future__ import annotations

import csv
import io
import json
from typing import Any, Literal

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel, Field

from app.research.sma_optimizer_service import sma_optimizer_service

router = APIRouter(prefix="/research/sma-optimizer", tags=["research"])

SORT_KEYS = (
    "score", "profit_factor", "expected_value", "win_rate",
    "net_points", "total_profit", "max_drawdown_points", "total_trades",
)


class RangeFloat(BaseModel):
    start: float
    end: float
    step: float


class RangeInt(BaseModel):
    start: int
    end: int
    step: int


class SmaOptimizerStartRequest(BaseModel):
    symbol: str = "SOLUSDT"
    timeframe: str = "5m"
    months_back: int = Field(default=6, ge=1, le=36)
    sma: RangeInt = Field(default_factory=lambda: RangeInt(start=20, end=200, step=2))
    stop: RangeFloat = Field(default_factory=lambda: RangeFloat(start=0.5, end=10.0, step=0.5))
    target: RangeFloat = Field(default_factory=lambda: RangeFloat(start=1.0, end=20.0, step=0.5))
    ambiguous: Literal["STOP_FIRST", "TARGET_FIRST", "IGNORE"] = "STOP_FIRST"


def _request_dict(body: SmaOptimizerStartRequest) -> dict[str, Any]:
    return {
        "symbol": body.symbol.upper(),
        "timeframe": body.timeframe,
        "months_back": body.months_back,
        "sma_start": body.sma.start,
        "sma_end": body.sma.end,
        "sma_step": body.sma.step,
        "stop_start": body.stop.start,
        "stop_end": body.stop.end,
        "stop_step": body.stop.step,
        "target_start": body.target.start,
        "target_end": body.target.end,
        "target_step": body.target.step,
        "ambiguous": body.ambiguous,
    }


@router.post("/preview")
def preview_grid(body: SmaOptimizerStartRequest) -> dict[str, Any]:
    return sma_optimizer_service.preview(_request_dict(body))


@router.post("/start")
def start_optimization(body: SmaOptimizerStartRequest) -> dict[str, Any]:
    req = _request_dict(body)
    plan = sma_optimizer_service.preview(req)
    total = plan["final_combinations"]
    if total <= 0:
        raise HTTPException(status_code=400, detail="No combinations in ranges")
    if total > 100_000:
        raise HTTPException(status_code=400, detail=f"Too many combinations ({total}). Narrow ranges.")
    return sma_optimizer_service.start(req)


@router.post("/stop/{job_id}")
def stop_optimization(job_id: str) -> dict[str, Any]:
    if not sma_optimizer_service.stop(job_id):
        raise HTTPException(status_code=404, detail="Job not found")
    return {"job_id": job_id, "status": "stopped"}


@router.get("/progress/{job_id}")
def optimization_progress(job_id: str) -> dict[str, Any]:
    progress = sma_optimizer_service.get_progress(job_id)
    if not progress:
        raise HTTPException(status_code=404, detail="Job not found")
    return progress


@router.get("/results/{job_id}")
def optimization_results(
    job_id: str,
    sort_by: str = Query(default="score"),
    include_trades: bool = Query(default=False),
) -> dict[str, Any]:
    if sort_by not in SORT_KEYS:
        raise HTTPException(status_code=400, detail=f"Invalid sort_by. Use one of: {SORT_KEYS}")
    payload = sma_optimizer_service.get_results(job_id, sort_by=sort_by, include_trades=include_trades)
    if not payload:
        raise HTTPException(status_code=404, detail="Job not found")
    return payload


@router.get("/results/{job_id}/trades/{result_index}")
def result_trades(
    job_id: str,
    result_index: int,
    sort_by: str = Query(default="score"),
) -> dict[str, Any]:
    payload = sma_optimizer_service.get_trades(job_id, result_index, sort_by=sort_by)
    if not payload:
        raise HTTPException(status_code=404, detail="Job or result not found")
    return payload


@router.get("/heatmap/{job_id}")
def heatmap(
    job_id: str,
    chart: str = Query(...),
    metric: str = Query(default="win_rate"),
    fix_sma: int | None = None,
    fix_stop: float | None = None,
    fix_target: float | None = None,
) -> dict[str, Any]:
    data = sma_optimizer_service.heatmap(
        job_id,
        chart=chart,
        fix_sma=fix_sma,
        fix_stop=fix_stop,
        fix_target=fix_target,
        metric=metric,
    )
    if not data:
        raise HTTPException(status_code=404, detail="Job not found or invalid chart")
    return data


_TOP_FIELDS = [
    "rank", "sma_length", "target_points", "stop_points", "total_trades",
    "win_rate", "profit_factor", "expected_value", "net_points", "score",
]

_RESULT_FIELDS = [
    "sma_length", "stop_points", "target_points", "total_trades",
    "winning_trades", "losing_trades", "win_rate", "avg_win", "avg_loss",
    "avg_bars_to_target", "avg_bars_to_stop", "profit_factor",
    "expected_value", "net_points", "total_profit", "max_drawdown_points",
    "max_consecutive_wins", "max_consecutive_losses", "largest_win",
    "largest_loss", "avg_duration_seconds", "score",
]


@router.get("/export/{job_id}/csv")
def export_csv(job_id: str, sort_by: str = Query(default="score")) -> StreamingResponse:
    payload = sma_optimizer_service.get_results(job_id, sort_by=sort_by)
    if not payload:
        raise HTTPException(status_code=404, detail="Job not found")
    rows = payload.get("results") or []
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=_RESULT_FIELDS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="sma-optimizer-{job_id}.csv"'},
    )


@router.get("/export/{job_id}/top-csv")
def export_top_csv(job_id: str, sort_by: str = Query(default="score")) -> StreamingResponse:
    payload = sma_optimizer_service.get_results(job_id, sort_by=sort_by)
    if not payload:
        raise HTTPException(status_code=404, detail="Job not found")
    rows = payload.get("top_results") or []
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=_TOP_FIELDS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="sma-optimizer-top20-{job_id}.csv"'},
    )


@router.get("/export/{job_id}/xlsx")
def export_xlsx(job_id: str, sort_by: str = Query(default="score")) -> Response:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=501,
            detail="Excel export requires openpyxl on the server",
        ) from exc

    payload = sma_optimizer_service.get_results(job_id, sort_by=sort_by, include_trades=True)
    if not payload:
        raise HTTPException(status_code=404, detail="Job not found")

    wb = Workbook()
    ws_top = wb.active
    ws_top.title = "Top20"
    ws_top.append(_TOP_FIELDS)
    for row in payload.get("top_results") or []:
        ws_top.append([row.get(f) for f in _TOP_FIELDS])

    ws_all = wb.create_sheet("AllResults")
    ws_all.append(_RESULT_FIELDS)
    for row in payload.get("results") or []:
        ws_all.append([row.get(f) for f in _RESULT_FIELDS])

    ws_trades = wb.create_sheet("Trades")
    trade_fields = [
        "sma_length", "stop_points", "target_points", "entry_time", "direction",
        "entry", "exit_price", "target", "stop", "bars", "result", "pnl_points",
    ]
    ws_trades.append(trade_fields)
    for row in payload.get("results") or []:
        base = {
            "sma_length": row.get("sma_length"),
            "stop_points": row.get("stop_points"),
            "target_points": row.get("target_points"),
        }
        for trade in row.get("trades") or []:
            ws_trades.append([base.get("sma_length"), base.get("stop_points"), base.get("target_points"),
                              trade.get("entry_time"), trade.get("direction"), trade.get("entry"),
                              trade.get("exit_price"), trade.get("target"), trade.get("stop"),
                              trade.get("bars"), trade.get("result"), trade.get("pnl_points")])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="sma-optimizer-{job_id}.xlsx"'},
    )


@router.get("/export/{job_id}/json")
def export_json(job_id: str, sort_by: str = Query(default="score")) -> Response:
    payload = sma_optimizer_service.get_results(job_id, sort_by=sort_by, include_trades=True)
    if not payload:
        raise HTTPException(status_code=404, detail="Job not found")
    return Response(
        content=json.dumps(payload, indent=2),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="sma-optimizer-{job_id}.json"'},
    )
